from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import F, Sum, Case, When, Value, DecimalField
from main.decorators import login_required_session, role_required
from main.models import User, ServiceStation, Booking, CarHistory
from main.views import get_current_user
from main.pdf_utils import generate_financial_report_pdf
from .models import Employee, SalaryBalance, Transaction, SparePart, UsedSparePart
from .supplier_api import search_supplier_parts, SUPPLIERS
import datetime
import re
from decimal import Decimal, InvalidOperation
import calendar
import csv
import json
import logging

logger = logging.getLogger(__name__)


def _redirect_to_dashboard(station_pk):
    return redirect(reverse('accounting:dashboard') + f'?station_id={station_pk}')


def _parse_date_range(request):
    today = datetime.date.today()
    first_day = today.replace(day=1)
    _, last_day_num = calendar.monthrange(today.year, today.month)
    last_day = today.replace(day=last_day_num)

    start_date = first_day
    end_date = last_day

    if start_str := request.GET.get('start_date'):
        try:
            start_date = datetime.datetime.strptime(start_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_str := request.GET.get('end_date'):
        try:
            end_date = datetime.datetime.strptime(end_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    return start_date, end_date



@login_required_session
@role_required('station')
def dashboard_view(request):
    user = get_current_user(request)
    if not user:
        return redirect('login')

    stations = ServiceStation.objects.filter(user=user)
    if not stations.exists():
        messages.warning(request, "Будь ласка, спочатку створіть СТО в профілі.")
        return redirect('profile')

    # Визначаємо, для якої саме СТО показувати звітність
    station_id = request.GET.get('station_id')
    selected_station = None
    if station_id:
        try:
            selected_station = stations.filter(pk=int(station_id)).first()
        except ValueError:
            pass
    if not selected_station:
        selected_station = stations.first()

    # Отримуємо діапазон дат для звітів
    start_date, end_date = _parse_date_range(request)

    # Отримуємо фільтри з GET-запиту
    t_type = request.GET.get('type', '')
    category = request.GET.get('category', '')
    employee_id_filter = request.GET.get('employee_id', '')

    # Працівники цієї станції
    employees = Employee.objects.filter(station=selected_station).select_related('salary_balance')

    # Отримуємо всі фінансові записи цієї станції
    all_transactions = Transaction.objects.filter(station=selected_station).select_related('employee', 'booking')
    
    # Фільтруємо транзакції за обраний період часу (базовий список транзакцій для метрик і графіків)
    period_transactions = all_transactions.filter(date__range=[start_date, end_date])

    # Розрахунок загальних фінансових показників через SQL-агрегацію
    totals = period_transactions.aggregate(
        total_income=Sum(
            Case(
                When(type='income', then='amount'),
                default=Value(Decimal('0.00')),
                output_field=DecimalField(),
            )
        ),
        total_expense=Sum(
            Case(
                When(type='expense', then='amount'),
                default=Value(Decimal('0.00')),
                output_field=DecimalField(),
            )
        ),
    )
    total_income = totals['total_income'] or Decimal('0.00')
    total_expense = totals['total_expense'] or Decimal('0.00')
    net_profit = total_income - total_expense

    # Формуємо дані для лінійного графіка динаміки фінансів за днями
    daily_data = {}
    curr_d = start_date
    while curr_d <= end_date:
        daily_data[curr_d.strftime("%d.%m")] = {'income': Decimal('0.00'), 'expense': Decimal('0.00')}
        curr_d += datetime.timedelta(days=1)

    for t in period_transactions:
        t_date_str = t.date.strftime("%d.%m")
        if t_date_str in daily_data:
            if t.type == 'income':
                daily_data[t_date_str]['income'] += t.amount
            else:
                daily_data[t_date_str]['expense'] += t.amount

    chart_dates = list(daily_data.keys())
    chart_incomes = [float(daily_data[d]['income']) for d in chart_dates]
    chart_expenses = [float(daily_data[d]['expense']) for d in chart_dates]

    # Групуємо витрати за категоріями для побудови діаграми
    expense_categories = {}
    for cat_code, cat_name in Transaction.TRANSACTION_CATEGORIES:
        if cat_code in ['salary', 'spare_parts', 'rent', 'utilities', 'other_expense']:
            expense_categories[cat_code] = {
                'label': cat_name.split(' (')[0],
                'amount': Decimal('0.00'),
                'percent': 0
            }

    period_expenses = [t for t in period_transactions if t.type == 'expense']
    for t in period_expenses:
        if t.category in expense_categories:
            expense_categories[t.category]['amount'] += t.amount
        else:
            if 'other_expense' in expense_categories:
                expense_categories['other_expense']['amount'] += t.amount

    if total_expense > 0:
        for cat in expense_categories:
            expense_categories[cat]['percent'] = int(
                round((expense_categories[cat]['amount'] / total_expense) * 100)
            )

    # Дані для кругової діаграми
    category_labels = []
    category_values = []
    for cat_code, cat_data in expense_categories.items():
        if cat_data['amount'] > 0:
            category_labels.append(cat_data['label'])
            category_values.append(float(cat_data['amount']))

    # Застосовуємо розширені фільтри до списку транзакцій, що відображається в таблиці
    filtered_transactions = period_transactions
    if t_type and t_type in ['income', 'expense']:
        filtered_transactions = filtered_transactions.filter(type=t_type)
    if category and category != 'all':
        filtered_transactions = filtered_transactions.filter(category=category)
    if employee_id_filter and employee_id_filter != 'all':
        try:
            filtered_transactions = filtered_transactions.filter(employee_id=int(employee_id_filter))
        except ValueError:
            pass

    # Окремий запис для виплати заробітних плат (архів виплат)
    salary_payouts = all_transactions.filter(category='salary', date__range=[start_date, end_date])
    if employee_id_filter and employee_id_filter != 'all':
        try:
            salary_payouts = salary_payouts.filter(employee_id=int(employee_id_filter))
        except ValueError:
            pass

    # Сума невиплачених зарплат активним працівникам
    unpaid_salaries = sum(
        sb.current_balance 
        for sb in SalaryBalance.objects.filter(
            employee__station=selected_station, 
            employee__is_active=True
        )
    )

    categories_choices = Transaction.TRANSACTION_CATEGORIES

    context = {
        'user': user,
        'stations': stations,
        'selected_station': selected_station,
        'employees': employees,
        'transactions': filtered_transactions[:100],  # Показуємо до 100 відфільтрованих записів
        'salary_payouts': salary_payouts[:100],        # До 100 записів архіву виплат
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'unpaid_salaries': unpaid_salaries,
        'start_date': start_date.strftime("%Y-%m-%d"),
        'end_date': end_date.strftime("%Y-%m-%d"),
        'categories_choices': categories_choices,
        'expense_categories': expense_categories.values(),
        
        # Параметри фільтрації для збереження стану полів
        'selected_type': t_type,
        'selected_category': category,
        'selected_employee_id': employee_id_filter,
        
        # Дані для передачі у скрипт Chart.js
        'chart_dates_json': json.dumps(chart_dates),
        'chart_incomes_json': json.dumps(chart_incomes),
        'chart_expenses_json': json.dumps(chart_expenses),
        'category_labels_json': json.dumps(category_labels),
        'category_values_json': json.dumps(category_values),
    }

    return render(request, 'accounting/dashboard.html', context)

@login_required_session
@role_required('station')
@require_POST
def add_employee_view(request):
    user = get_current_user(request)
    station_id = request.POST.get('station_id')
    station = get_object_or_404(ServiceStation, pk=station_id, user=user)

    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    position = request.POST.get('position', '').strip()
    base_salary = request.POST.get('base_salary', '0.00')
    commission_percent = request.POST.get('commission_percent', '0.00')

    if not full_name or not position:
        messages.error(request, "Ім'я та посада є обов'язковими.")
        return _redirect_to_dashboard(station.pk)

    try:
        Employee.objects.create(
            station=station,
            full_name=full_name,
            phone=phone if phone else None,
            email=email if email else None,
            position=position,
            base_salary=Decimal(base_salary),
            commission_percent=Decimal(commission_percent)
        )
        messages.success(request, f"Працівника {full_name} успішно додано.")
    except (InvalidOperation, Exception):
        logger.error("Помилка при додаванні працівника", exc_info=True)
        messages.error(request, "Помилка при додаванні працівника. Спробуйте пізніше.")

    return _redirect_to_dashboard(station.pk)

@login_required_session
@role_required('station')
@require_POST
def edit_employee_view(request, employee_id):
    user = get_current_user(request)
    employee = get_object_or_404(Employee, pk=employee_id, station__user=user)

    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    email = request.POST.get('email', '').strip()
    position = request.POST.get('position', '').strip()
    base_salary = request.POST.get('base_salary', '0.00')
    commission_percent = request.POST.get('commission_percent', '0.00')
    is_active = request.POST.get('is_active') == 'true'

    if not full_name or not position:
        messages.error(request, "Ім'я та посада є обов'язковими.")
        return _redirect_to_dashboard(employee.station.pk)

    try:
        employee.full_name = full_name
        employee.phone = phone if phone else None
        employee.email = email if email else None
        employee.position = position
        employee.base_salary = Decimal(base_salary)
        employee.commission_percent = Decimal(commission_percent)
        employee.is_active = is_active
        employee.save()
        messages.success(request, f"Дані працівника {full_name} успішно оновлено.")
    except (InvalidOperation, Exception):
        logger.error("Помилка при оновленні працівника", exc_info=True)
        messages.error(request, "Помилка при оновленні даних працівника. Спробуйте пізніше.")

    return _redirect_to_dashboard(employee.station.pk)

@login_required_session
@role_required('station')
@require_POST
def fire_employee_view(request, employee_id):
    user = get_current_user(request)
    employee = get_object_or_404(Employee, pk=employee_id, station__user=user)
    
    employee.is_active = False
    employee.save()
    messages.success(request, f"Працівника {employee.full_name} звільнено.")
    return _redirect_to_dashboard(employee.station.pk)

@login_required_session
@role_required('station')
@require_POST
def pay_salary_view(request):
    """
    Виплата заробітної плати працівнику СТО.
    Транзакція виконується атомарно з блокуванням рядка для уникнення race condition.
    """
    user = get_current_user(request)
    employee_id = request.POST.get('employee_id')
    employee = get_object_or_404(Employee, pk=employee_id, station__user=user)

    amount_str = request.POST.get('amount', '0.00')
    try:
        amount = Decimal(amount_str)
    except (ValueError, InvalidOperation):
        messages.error(request, "Некоректна сума виплати.")
        return _redirect_to_dashboard(employee.station.pk)

    if amount <= 0:
        messages.error(request, "Сума виплати має бути більшою за нуль.")
        return _redirect_to_dashboard(employee.station.pk)

    try:
        with transaction.atomic():
            # Блокуємо рядок балансу у БД від паралельних змін
            balance = SalaryBalance.objects.select_for_update().get(
                employee=employee
            )
            if amount > balance.current_balance:
                messages.error(
                    request,
                    f"Сума виплати ({amount} грн) перевищує "
                    f"доступний баланс ({balance.current_balance} грн)."
                )
                return _redirect_to_dashboard(employee.station.pk)

            # Атомарне оновлення значення суми виплат безпосередньо в базі
            balance.total_paid = F('total_paid') + amount
            balance.save(update_fields=['total_paid'])

            Transaction.objects.create(
                station=employee.station,
                type='expense',
                category='salary',
                amount=amount,
                description=(
                    f"Виплата зарплати працівнику: "
                    f"{employee.full_name} ({employee.position})"
                ),
                employee=employee,
                date=datetime.date.today()
            )

        messages.success(
            request,
            f"Виплата {amount} грн працівнику {employee.full_name} "
            f"успішно проведена."
        )
    except Exception:
        logger.error("Помилка при виплаті зарплати", exc_info=True)
        messages.error(request, "Помилка при виплаті зарплати. Спробуйте пізніше.")

    return _redirect_to_dashboard(employee.station.pk)

@login_required_session
@role_required('station')
@require_POST
def add_transaction_view(request):
    user = get_current_user(request)
    station_id = request.POST.get('station_id')
    station = get_object_or_404(ServiceStation, pk=station_id, user=user)

    t_type = request.POST.get('type')
    category = request.POST.get('category')
    amount_str = request.POST.get('amount')
    description = request.POST.get('description', '').strip()
    date_str = request.POST.get('date')

    # Валідація типу
    if t_type not in ['income', 'expense']:
        messages.error(request, "Невірний тип операції.")
        return _redirect_to_dashboard(station.pk)

    # Валідація категорії по списку допустимих значень
    valid_categories = dict(Transaction.TRANSACTION_CATEGORIES)
    if not category or category not in valid_categories:
        messages.error(request, "Невірна категорія операції.")
        return _redirect_to_dashboard(station.pk)

    if not amount_str:
        messages.error(request, "Заповніть суму операції.")
        return _redirect_to_dashboard(station.pk)

    try:
        amount = Decimal(amount_str)
    except (ValueError, InvalidOperation):
        messages.error(request, "Некоректна сума операції.")
        return _redirect_to_dashboard(station.pk)

    if amount <= 0:
        messages.error(request, "Сума має бути більшою за нуль.")
        return _redirect_to_dashboard(station.pk)

    try:
        t_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.date.today()

        Transaction.objects.create(
            station=station,
            type=t_type,
            category=category,
            amount=amount,
            description=description,
            date=t_date
        )
        messages.success(request, "Операцію успішно додано.")
    except Exception:
        logger.error("Помилка при додаванні операції", exc_info=True)
        messages.error(request, "Помилка при додаванні операції. Спробуйте пізніше.")

    return _redirect_to_dashboard(station.pk)

@login_required_session
@role_required('station')
@require_POST
def complete_booking_view(request):
    """
    Завершення замовлення на ремонт СТО.
    Встановлює відповідний статус, фіксує дохід та нараховує комісію майстру.
    """
    user = get_current_user(request)
    booking_id = request.POST.get('booking_id')

    actual_price_str = request.POST.get('actual_price')
    employee_id = request.POST.get('employee_id')
    mileage_str = request.POST.get('mileage')
    work_list = request.POST.get('work_list')
    spare_parts = request.POST.get('spare_parts')
    used_parts_json = request.POST.get('used_parts_json')

    mileage = None
    if mileage_str:
        try:
            mileage = int(mileage_str)
        except ValueError:
            pass

    used_parts_data = []
    if used_parts_json:
        try:
            used_parts_data = json.loads(used_parts_json)
        except Exception:
            used_parts_data = []

    if not actual_price_str:
        messages.error(
            request,
            "Будь ласка, вкажіть фактичну вартість ремонту."
        )
        return redirect('profile')

    try:
        actual_price = Decimal(actual_price_str)
    except (ValueError, InvalidOperation):
        messages.error(request, "Некоректна сума вартості ремонту.")
        return redirect('profile')

    if actual_price <= 0:
        messages.error(
            request,
            "Вартість ремонту повинна бути більшою за нуль."
        )
        return redirect('profile')

    try:
        with transaction.atomic():
            # Блокуємо рядок замовлення від паралельних змін іншими процесами
            booking = Booking.objects.select_for_update().get(
                pk=booking_id, station__user=user
            )

            # Перевірка поточного статусу
            if booking.status == 'completed':
                messages.warning(
                    request,
                    f"Заявка #{booking.pk} вже була завершена раніше."
                )
                return redirect(reverse('profile') + '?tab=bookings')

            if booking.status == 'cancelled':
                messages.error(
                    request,
                    f"Неможливо завершити скасовану заявку #{booking.pk}."
                )
                return redirect(reverse('profile') + '?tab=bookings')

            employee = None
            if employee_id:
                employee = get_object_or_404(
                    Employee, pk=employee_id, station=booking.station
                )

            # 1. Позначаємо ремонт як виконаний
            booking.status = 'completed'
            booking.save(update_fields=['status'])

            # 2. Записуємо вартість робіт у доходи СТО
            desc = (
                f"Завершено ремонт за заявкою #{booking.id} "
                f"({booking.service_name or 'Загальні роботи'})"
            )
            if employee:
                desc += f". Виконавець: {employee.full_name}."

            tx = Transaction.objects.create(
                station=booking.station,
                type='income',
                category='service',
                amount=actual_price,
                description=desc,
                booking=booking,
                employee=employee,
                date=datetime.date.today()
            )

            # 3. Нараховуємо майстру комісію
            if employee and employee.commission_percent > 0:
                commission = actual_price * (
                    employee.commission_percent / Decimal('100.00')
                )
                commission = commission.quantize(Decimal('0.01'))

                # Блокуємо та оновлюємо баланс заробітної плати майстра
                sb = SalaryBalance.objects.select_for_update().get(
                    employee=employee
                )
                sb.total_earned = F('total_earned') + commission
                sb.save(update_fields=['total_earned'])

                tx.description += f" Нараховано комісію: {commission} грн."
                tx.save(update_fields=['description'])

            # 4. Обробка вибраних запчастин зі складу (списання та обчислення собівартості)
            total_parts_cost = Decimal('0.00')
            parts_summary_lines = []

            for item in used_parts_data:
                p_id = item.get('part_id')
                p_qty = int(item.get('qty', 0))
                if not p_id or p_qty <= 0:
                    continue
                try:
                    sp = SparePart.objects.select_for_update().get(pk=p_id, station=booking.station)
                    actual_deduct = min(sp.quantity, p_qty) if sp.quantity > 0 else 0
                    if actual_deduct > 0:
                        sp.quantity = F('quantity') - actual_deduct
                        sp.save(update_fields=['quantity'])
                        sp.refresh_from_db()

                        UsedSparePart.objects.create(
                            booking=booking,
                            spare_part=sp,
                            part_name=sp.name,
                            quantity=actual_deduct,
                            cost_price=sp.cost_price,
                            selling_price=sp.selling_price
                        )

                        line_cost = sp.cost_price * actual_deduct
                        line_sell = sp.selling_price * actual_deduct
                        total_parts_cost += line_cost
                        parts_summary_lines.append(f"{sp.name} x{actual_deduct} ({line_sell} грн)")
                except SparePart.DoesNotExist:
                    continue

            # Фіксуємо собівартість запчастин у витратах СТО
            if total_parts_cost > 0:
                Transaction.objects.create(
                    station=booking.station,
                    type='expense',
                    category='spare_parts',
                    amount=total_parts_cost,
                    description=f"Списання собівартості запчастин за заявкою #{booking.id}",
                    booking=booking,
                    employee=employee,
                    date=datetime.date.today()
                )

            # 5. Автоматично створюємо запис в історії обслуговування авто
            if parts_summary_lines:
                auto_parts_str = ", ".join(parts_summary_lines)
                if spare_parts and spare_parts.strip():
                    spare_parts = f"{auto_parts_str}\n{spare_parts.strip()}"
                else:
                    spare_parts = auto_parts_str

            if booking.car:
                final_works = work_list.strip() if (work_list and work_list.strip()) else (booking.service_name or booking.description or 'Виконано роботи з обслуговування')
                CarHistory.objects.create(
                    car=booking.car,
                    booking=booking,
                    station=booking.station,
                    date=datetime.date.today(),
                    mileage=mileage,
                    work_list=final_works,
                    spare_parts=spare_parts.strip() if spare_parts else '',
                    price=actual_price
                )

        messages.success(
            request,
            f"Ремонт за заявкою #{booking.id} успішно завершено. "
            f"Суму {actual_price} грн внесено в дохід СТО, а деталі списано зі складу."
        )
    except Booking.DoesNotExist:
        messages.error(request, "Заявку не знайдено.")
    except Exception:
        logger.error("Помилка при завершенні ремонту", exc_info=True)
        messages.error(request, "Помилка при завершенні ремонту. Спробуйте пізніше.")

    return redirect(reverse('profile') + '?tab=bookings')


@login_required_session
@role_required('station')
@require_POST
def add_spare_part_view(request):
    """
    Додавання нової запчастини на склад СТО.
    """
    user = get_current_user(request)
    station_id = request.POST.get('station_id')
    station = get_object_or_404(ServiceStation, pk=station_id, user=user)

    name = request.POST.get('name', '').strip()
    sku = request.POST.get('sku', '').strip()
    quantity_str = request.POST.get('quantity', '0')
    cost_price_str = request.POST.get('cost_price', '0.00')
    selling_price_str = request.POST.get('selling_price', '0.00')
    min_quantity_str = request.POST.get('min_quantity', '5')

    if not name:
        messages.error(request, "Будь ласка, вкажіть назву запчастини.")
        return _redirect_to_dashboard(station.pk)

    try:
        quantity = max(0, int(quantity_str))
        cost_price = max(Decimal('0.00'), Decimal(cost_price_str))
        selling_price = max(Decimal('0.00'), Decimal(selling_price_str))
        min_quantity = max(0, int(min_quantity_str))
    except (ValueError, InvalidOperation):
        messages.error(request, "Некоректні числові дані для запчастини.")
        return _redirect_to_dashboard(station.pk)

    SparePart.objects.create(
        station=station,
        name=name,
        sku=sku if sku else None,
        quantity=quantity,
        cost_price=cost_price,
        selling_price=selling_price,
        min_quantity=min_quantity
    )

    messages.success(request, f"Запчастину '{name}' успішно додано на склад.")
    return redirect(reverse('profile') + '?tab=inventory')


@login_required_session
@role_required('station')
@require_POST
def edit_spare_part_view(request):
    """
    Редагування складських залишків та цін запчастини.
    """
    user = get_current_user(request)
    part_id = request.POST.get('part_id')
    part = get_object_or_404(SparePart, pk=part_id, station__user=user)

    name = request.POST.get('name', '').strip()
    sku = request.POST.get('sku', '').strip()
    quantity_str = request.POST.get('quantity')
    cost_price_str = request.POST.get('cost_price')
    selling_price_str = request.POST.get('selling_price')
    min_quantity_str = request.POST.get('min_quantity')

    if name:
        part.name = name
    part.sku = sku if sku else None

    try:
        if quantity_str is not None and quantity_str != '':
            part.quantity = max(0, int(quantity_str))
        if cost_price_str is not None and cost_price_str != '':
            part.cost_price = max(Decimal('0.00'), Decimal(cost_price_str))
        if selling_price_str is not None and selling_price_str != '':
            part.selling_price = max(Decimal('0.00'), Decimal(selling_price_str))
        if min_quantity_str is not None and min_quantity_str != '':
            part.min_quantity = max(0, int(min_quantity_str))
        part.save()
        messages.success(request, f"Запчастину '{part.name}' оновлено.")
    except (ValueError, InvalidOperation):
        messages.error(request, "Помилка оновлення даних запчастини.")

    return redirect(reverse('profile') + '?tab=inventory')


@login_required_session
@role_required('station')
@require_POST
def delete_spare_part_view(request):
    """
    Видалення позиції запчастини зі складу.
    """
    user = get_current_user(request)
    part_id = request.POST.get('part_id')
    part = get_object_or_404(SparePart, pk=part_id, station__user=user)
    name = part.name
    part.delete()
    messages.success(request, f"Запчастину '{name}' видалено зі складу.")
    return redirect(reverse('profile') + '?tab=inventory')


@login_required_session
@role_required('station')
def export_transactions_csv(request):
    """
    Експорт фінансових транзакцій СТО у формат CSV.
    """
    user = get_current_user(request)
    if not user:
        return redirect('login')

    station_id = request.GET.get('station_id')
    if not station_id:
        messages.error(request, "Не вказано СТО для експорту.")
        return redirect('profile')

    station = get_object_or_404(ServiceStation, pk=int(station_id), user=user)

    # Отримуємо обраний діапазон дат з утиліти
    start_date, end_date = _parse_date_range(request)

    t_type = request.GET.get('type')
    category = request.GET.get('category')
    employee_id_filter = request.GET.get('employee_id')

    transactions = Transaction.objects.filter(
        station=station, date__range=[start_date, end_date]
    ).select_related('employee', 'booking')

    if t_type and t_type in ['income', 'expense']:
        transactions = transactions.filter(type=t_type)
    if category and category != 'all':
        transactions = transactions.filter(category=category)
    if employee_id_filter and employee_id_filter != 'all':
        try:
            transactions = transactions.filter(
                employee_id=int(employee_id_filter)
            )
        except ValueError:
            pass

    # Видаляємо не-ASCII символи для сумісності з усіма браузерами
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '', re.sub(r'\s+', '_', station.name)).strip('_')[:50] or 'station'
    filename = f"{safe_name}_report_{start_date}_{end_date}.xlsx"

    # Створюємо книжку Excel з ошатним оформленням
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Фінансовий звіт"
    ws.views.sheetView[0].showGridLines = True

    # 1. Шапка документа
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"ФІНАНСОВИЙ ЗВІТ СТО: {station.name.upper()}"
    title_cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:G2')
    sub_cell = ws['A2']
    sub_cell.value = f"Період: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
    sub_cell.font = Font(name='Arial', size=10, italic=True, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 2. Заголовки таблиці
    headers = ['Дата', 'Тип операції', 'Категорія', 'Сума (грн)', 'Опис', 'ID Заявки', 'Співробітник']
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='medium', color='0284C7'),
        bottom=Side(style='medium', color='0284C7')
    )

    ws.append([]) # Порожній рядок 3
    ws.append(headers) # Рядок 4
    ws.row_dimensions[4].height = 26

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border

    # 3. Данні транзакцій
    row_fill_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    total_inc = Decimal('0.00')
    total_exp = Decimal('0.00')

    current_row = 5
    for t in transactions:
        row_data = [
            t.date.strftime("%d.%m.%Y"),
            t.get_type_display(),
            t.get_category_display(),
            float(t.amount),
            t.description or '',
            t.booking.id if t.booking else '',
            t.employee.full_name if t.employee else ''
        ]
        ws.append(row_data)
        ws.row_dimensions[current_row].height = 20
        fill = row_fill_even if current_row % 2 == 0 else row_fill_odd

        for c_idx in range(1, 8):
            cell = ws.cell(row=current_row, column=c_idx)
            cell.fill = fill
            cell.border = data_border
            cell.font = Font(name='Arial', size=10)

            # Вирівнювання та колір суми
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 4:
                cell.number_format = '#,##0.00 "грн"'
                cell.alignment = Alignment(horizontal="right")
                if t.type == 'income':
                    cell.font = Font(name='Arial', size=10, bold=True, color="16A34A")
                    total_inc += t.amount
                else:
                    cell.font = Font(name='Arial', size=10, bold=True, color="DC2626")
                    total_exp += t.amount

        current_row += 1

    # 4. Підсумковий рядок
    ws.append([])
    summary_row = current_row + 1
    ws.cell(row=summary_row, column=3, value="ЧИСТИЙ ПРИБУТОК:").font = Font(name='Arial', size=11, bold=True, color="0F172A")
    ws.cell(row=summary_row, column=3).alignment = Alignment(horizontal="right")
    
    net_cell = ws.cell(row=summary_row, column=4, value=float(total_inc - total_exp))
    net_cell.font = Font(name='Arial', size=12, bold=True, color="16A34A" if total_inc >= total_exp else "DC2626")
    net_cell.number_format = '#,##0.00 "грн"'
    net_cell.border = Border(top=Side(style='thin', color='0F172A'), bottom=Side(style='double', color='0F172A'))

    # Авто-налаштування ширини колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: # Пропускаємо об'єднані заголовки
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required_session
@role_required('station')
def export_financial_report_pdf(request):
    """
    Формує та повертає Фінансовий звіт СТО за обраний період у форматі PDF.
    """
    user = get_current_user(request)
    if not user:
        return redirect('login')

    stations = ServiceStation.objects.filter(user=user)
    if not stations.exists():
        messages.warning(request, "Будь ласка, спочатку створіть СТО.")
        return redirect('profile')

    station_id = request.GET.get('station_id')
    selected_station = None
    if station_id:
        try:
            selected_station = stations.filter(pk=int(station_id)).first()
        except ValueError:
            pass
    if not selected_station:
        selected_station = stations.first()

    start_date, end_date = _parse_date_range(request)

    # Отримуємо всі операції за період
    all_transactions = Transaction.objects.filter(station=selected_station).select_related('employee', 'booking')
    period_transactions = list(all_transactions.filter(date__range=[start_date, end_date]))

    total_income = sum((t.amount for t in period_transactions if t.type == 'income'), Decimal('0.00'))
    total_expense = sum((t.amount for t in period_transactions if t.type == 'expense'), Decimal('0.00'))
    net_profit = total_income - total_expense
    profit_margin = float((net_profit / total_income) * 100) if total_income > 0 else 0.0

    completed_bookings = Booking.objects.filter(
        station=selected_station,
        status='completed',
        created_at__date__range=[start_date, end_date]
    ).count()

    # Формуємо розбивку за категоріями
    income_by_category = {}
    expense_by_category = {}

    for t in period_transactions:
        cat_disp = t.get_category_display()
        if t.type == 'income':
            income_by_category[cat_disp] = income_by_category.get(cat_disp, Decimal('0.00')) + t.amount
        else:
            expense_by_category[cat_disp] = expense_by_category.get(cat_disp, Decimal('0.00')) + t.amount

    metrics = {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'profit_margin': profit_margin,
        'completed_bookings': completed_bookings,
        'income_by_category': income_by_category,
        'expense_by_category': expense_by_category,
    }

    employees = list(Employee.objects.filter(station=selected_station).select_related('salary_balance'))

    # Генерація PDF-байтів
    pdf_bytes = generate_financial_report_pdf(
        selected_station, start_date, end_date, period_transactions, metrics, employees
    )

    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '', re.sub(r'\s+', '_', selected_station.name)).strip('_')[:50] or 'station'
    filename = f"financial_report_{safe_name}_{start_date}_{end_date}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    if request.GET.get('inline') == '1':
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    else:
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def search_supplier_parts_api(request):
    """
    API пошуку запчастин у каталогах постачальників (InterCars, Exist, TechnoVector).
    Приймає ?query=... та ?supplier=...
    """
    query = request.GET.get('query', '').strip()
    supplier = request.GET.get('supplier', 'all').strip()
    data = search_supplier_parts(query, supplier)
    return JsonResponse(data)


@require_POST
@login_required_session
@role_required('station')
def import_supplier_part_view(request):
    """
    Імпорт/замовлення обраної запчастини з каталогу постачальника на склад СТО.
    """
    user = get_current_user(request)
    station_id = request.POST.get('station_id')
    station = get_object_or_404(ServiceStation, pk=station_id, user=user)

    sku = request.POST.get('sku', '').strip()
    part_name = request.POST.get('part_name', '').strip()
    brand = request.POST.get('brand', '').strip()
    cost_price_str = request.POST.get('cost_price', '0')
    selling_price_str = request.POST.get('selling_price', '0')
    quantity_str = request.POST.get('quantity', '1')

    if not part_name:
        messages.error(request, 'Назва запчастини є обов\'язковою.')
        return _redirect_to_dashboard(station.pk)

    try:
        cost_price = Decimal(cost_price_str)
        selling_price = Decimal(selling_price_str)
        quantity = int(quantity_str)
    except (InvalidOperation, ValueError):
        messages.error(request, 'Невірний формат ціни або кількості.')
        return _redirect_to_dashboard(station.pk)

    full_name = f"{part_name} ({brand})" if brand else part_name

    existing_part = SparePart.objects.filter(station=station, name=full_name, sku=sku).first()
    if existing_part:
        existing_part.quantity += quantity
        existing_part.cost_price = cost_price
        if selling_price > 0:
            existing_part.selling_price = selling_price
        existing_part.save()
        messages.success(request, f'Кількість запчастини "{full_name}" оновлено на складі (+{quantity} шт).')
    else:
        if selling_price <= 0:
            selling_price = round(cost_price * Decimal('1.35'), 2)
        
        SparePart.objects.create(
            station=station,
            name=full_name,
            sku=sku,
            quantity=quantity,
            cost_price=cost_price,
            selling_price=selling_price,
            min_quantity=3
        )
        messages.success(request, f'Запчастину "{full_name}" закуплено та додано на склад СТО.')

    return _redirect_to_dashboard(station.pk)

